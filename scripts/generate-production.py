"""
Fetches EIA v2 API crude oil production data and writes two output files:

1. public/data/production-basins.json  — basin centroids + latest production (bbl/d)
2. public/data/production-history.json — 36-month time series per basin

Source: EIA v2 API (free key required)
  https://api.eia.gov/v2/petroleum/crd/drill/data/
  Register for a free API key at https://www.eia.gov/opendata/

Usage:
  EIA_API_KEY=your_key python3 scripts/generate-production.py

If EIA changes the endpoint or series IDs, the script exits non-zero and the
app falls back to whatever is committed in public/data/.

Note: If you don't have an EIA API key you can still run this by setting
EIA_API_KEY to an empty string — the script will fall back to the bundled
DPR xlsx if present at data/raw/dpr-data.xlsx.
"""

import json
import os
import sys
from datetime import datetime, date
from pathlib import Path

try:
    import requests
except ImportError:
    sys.exit("requests not installed. Run: pip install requests")

EIA_API_KEY = os.environ.get("EIA_API_KEY", "")
EIA_V2_URL = "https://api.eia.gov/v2/petroleum/crd/drill/data/"

DPR_FALLBACK = Path("data/raw/dpr-data.xlsx")
OUT_BASINS = Path("public/data/production-basins.json")
OUT_HISTORY = Path("public/data/production-history.json")

# EIA DPR region codes → display names and map metadata
BASIN_META = {
    "ANK": {
        "display": "Anadarko Basin",
        "lon": -98.5, "lat": 35.5,
        "states": ["OK", "KS", "TX"],
        "color": "#f59e0b",
    },
    "APP": {
        "display": "Appalachian Basin",
        "lon": -80.2, "lat": 40.0,
        "states": ["PA", "WV", "OH"],
        "color": "#10b981",
    },
    "BAK": {
        "display": "Bakken",
        "lon": -103.0, "lat": 47.5,
        "states": ["ND", "MT"],
        "color": "#3b82f6",
    },
    "EAF": {
        "display": "Eagle Ford",
        "lon": -98.5, "lat": 28.8,
        "states": ["TX"],
        "color": "#f97316",
    },
    "HAY": {
        "display": "Haynesville",
        "lon": -93.5, "lat": 32.2,
        "states": ["LA", "TX"],
        "color": "#8b5cf6",
    },
    "NIO": {
        "display": "Niobrara / DJ Basin",
        "lon": -104.5, "lat": 41.0,
        "states": ["CO", "WY"],
        "color": "#ec4899",
    },
    "PER": {
        "display": "Permian Basin",
        "lon": -102.0, "lat": 31.8,
        "states": ["TX", "NM"],
        "color": "#ef4444",
    },
}

# EIA v2 series IDs for total oil production per region (bbl/d)
# Series format: PET.XXXXXX.M  — fetched via the drillprod endpoint
EIA_SERIES = {
    "ANK": "PET.OEGR_ANK.M",
    "APP": "PET.OEGR_APP.M",
    "BAK": "PET.OEGR_BAK.M",
    "EAF": "PET.OEGR_EAF.M",
    "HAY": "PET.OEGR_HAY.M",
    "NIO": "PET.OEGR_NIO.M",
    "PER": "PET.OEGR_PER.M",
}


def fetch_eia_v2_series(series_id: str, months: int = 36) -> list[dict]:
    """Fetch a single EIA v2 time series. Returns [{month, bpd}, ...] sorted asc."""
    params = {
        "api_key": EIA_API_KEY,
        "frequency": "monthly",
        "data[0]": "value",
        "series_id": series_id,
        "sort[0][column]": "period",
        "sort[0][direction]": "desc",
        "length": months,
        "offset": 0,
    }
    resp = requests.get(EIA_V2_URL, params=params, timeout=30)
    print(f"  GET {resp.url[:120]}")
    resp.raise_for_status()
    payload = resp.json()

    if "response" not in payload:
        raise RuntimeError(f"Unexpected EIA response shape: {list(payload.keys())}")

    data = payload["response"].get("data", [])
    entries = []
    for item in data:
        period = item.get("period", "")       # YYYY-MM
        value = item.get("value")
        if not period or value is None:
            continue
        try:
            bpd = round(float(value) * 1000)  # EIA reports in thousand bbl/d
        except (ValueError, TypeError):
            continue
        entries.append({"month": period[:7], "bpd": bpd})

    entries.sort(key=lambda x: x["month"])
    return entries


def fetch_all_via_eia_v2() -> dict[str, list[dict]]:
    """Fetch 36-month history for all basins via EIA v2 API."""
    result = {}
    for code, series_id in EIA_SERIES.items():
        display = BASIN_META[code]["display"]
        print(f"  Fetching {display} ({series_id}) ...")
        try:
            entries = fetch_eia_v2_series(series_id)
            if entries:
                result[display] = entries
                print(f"    → {len(entries)} months ({entries[0]['month']} – {entries[-1]['month']})")
            else:
                print(f"    WARNING: no data returned for {series_id}")
        except Exception as e:
            print(f"    ERROR: {e}")
    return result


def fallback_from_xlsx() -> dict[str, list[dict]]:
    """Fall back to committed dpr-data.xlsx if EIA API is unavailable."""
    if not DPR_FALLBACK.exists():
        return {}
    print(f"Falling back to {DPR_FALLBACK} ...")
    try:
        import openpyxl
    except ImportError:
        print("openpyxl not installed — cannot use xlsx fallback")
        return {}

    XLSX_SHEET_MAP = {
        "Anadarko Region": "ANK",
        "Appalachia Region": "APP",
        "Bakken Region": "BAK",
        "Eagle Ford Region": "EAF",
        "Haynesville Region": "HAY",
        "Niobrara Region": "NIO",
        "Permian Region": "PER",
    }

    wb = openpyxl.load_workbook(str(DPR_FALLBACK))
    result = {}
    for sheet_name, code in XLSX_SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        monthly = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row[0] or not isinstance(row[0], datetime):
                continue
            oil = row[4]
            if oil is None:
                continue
            monthly.append({"month": row[0].strftime("%Y-%m"), "bpd": round(float(oil))})
        if monthly:
            display = BASIN_META[code]["display"]
            result[display] = monthly[-36:]
    return result


def build_outputs(history: dict[str, list[dict]]) -> tuple[list, dict]:
    basins = []
    for code, meta in BASIN_META.items():
        display = meta["display"]
        entries = history.get(display)
        if not entries:
            continue
        latest = entries[-1]
        basins.append({
            "id": code.lower(),
            "name": display,
            "lon": meta["lon"],
            "lat": meta["lat"],
            "states": meta["states"],
            "color": meta["color"],
            "bpd": latest["bpd"],
            "month": latest["month"],
        })
    basins.sort(key=lambda x: -x["bpd"])
    return basins, history


def main():
    if not EIA_API_KEY:
        print("EIA_API_KEY not set — trying xlsx fallback")
        history = fallback_from_xlsx()
        if not history:
            sys.exit(
                "No data source available.\n"
                "Set EIA_API_KEY (free at https://www.eia.gov/opendata/) "
                "or place dpr-data.xlsx at data/raw/dpr-data.xlsx"
            )
    else:
        print("Fetching production data from EIA v2 API ...")
        history = fetch_all_via_eia_v2()
        if not history:
            print("EIA v2 returned no data — falling back to xlsx")
            history = fallback_from_xlsx()
        if not history:
            sys.exit("All data sources failed")

    basins, history = build_outputs(history)

    OUT_BASINS.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_BASINS, "w") as f:
        json.dump(basins, f, indent=2)
    with open(OUT_HISTORY, "w") as f:
        json.dump(history, f)

    print(f"\nWrote {len(basins)} basins → {OUT_BASINS}")
    print(f"Wrote history ({len(history)} basins) → {OUT_HISTORY}")
    print()
    print(f"{'Basin':<25} {'Production (bbl/d)':>20} {'Month'}")
    print("-" * 55)
    for b in basins:
        print(f"{b['name']:<25} {b['bpd']:>20,} {b['month']}")


if __name__ == "__main__":
    main()
